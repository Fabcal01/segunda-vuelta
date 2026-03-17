import openpyxl
import json
import os

# ── Leer MAESTRO (DNIs) ──────────────────────────────────────────
wb = openpyxl.load_workbook('data/MAESTRO.xlsx')
ws = wb.active
users = {}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[0] and row[2]:
        users[str(row[0]).strip()] = str(row[2]).strip()

# ── Leer BD (Recorridos) ─────────────────────────────────────────
wb2 = openpyxl.load_workbook('data/BD.xlsx')
ws2 = wb2['Segunda Vuelta']
records = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    fecha = row[0]
    ot    = row[1]
    conductor = row[5]
    aux1  = row[7]
    aux2  = row[8]
    aux3  = row[9]
    if not fecha:
        continue
    personas = []
    for p, rol in [(conductor,'CONDUCTOR'),(aux1,'AUX1'),(aux2,'AUX2'),(aux3,'AUX3')]:
        if p:
            personas.append({'nombre': str(p).strip(), 'rol': rol})
    if personas:
        records.append({
            'fecha': fecha.strftime('%Y-%m-%d'),
            'ot': ot or '',
            'personas': personas
        })

# ── Detectar cortes disponibles (cada 21→20) ─────────────────────
from datetime import date, timedelta
fechas = sorted(set(r['fecha'] for r in records))
cortes = []
if fechas:
    f_min = date.fromisoformat(min(fechas))
    f_max = date.fromisoformat(max(fechas))
    # Generar cortes desde el primer mes hasta el último
    y, m = f_min.year, f_min.month
    while True:
        ini = date(y, m, 21)
        if m == 12:
            fin = date(y+1, 1, 20)
        else:
            fin = date(y, m+1, 20)
        if ini > f_max + timedelta(days=31):
            break
        label_ini = ini.strftime('%-d %b').replace('Jan','Ene').replace('Feb','Feb').replace('Mar','Mar').replace('Apr','Abr').replace('May','May').replace('Jun','Jun').replace('Jul','Jul').replace('Aug','Ago').replace('Sep','Sep').replace('Oct','Oct').replace('Nov','Nov').replace('Dec','Dic')
        label_fin = fin.strftime('%-d %b %Y').replace('Jan','Ene').replace('Feb','Feb').replace('Mar','Mar').replace('Apr','Abr').replace('May','May').replace('Jun','Jun').replace('Jul','Jul').replace('Aug','Ago').replace('Sep','Sep').replace('Oct','Oct').replace('Nov','Nov').replace('Dec','Dic')
        cortes.append({
            'value': f"{ini.isoformat()}|{fin.isoformat()}",
            'label': f"{label_ini} – {label_fin} {ini.year}" if ini.year == fin.year else f"{label_ini} {ini.year} – {label_fin}",
            'selected': ini <= f_max <= fin
        })
        m += 1
        if m > 12:
            m = 1
            y += 1

users_json   = json.dumps(users,   ensure_ascii=False)
records_json = json.dumps(records, ensure_ascii=False)
cortes_json  = json.dumps(cortes,  ensure_ascii=False)

# ── Leer template y generar index.html ───────────────────────────
with open('scripts/template.html', 'r', encoding='utf-8') as f:
    html = f.read()

html = html.replace('__USERS__',   users_json)
html = html.replace('__DATA__',    records_json)
html = html.replace('__CORTES__',  cortes_json)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f"✅ index.html generado: {len(users)} colaboradores, {len(records)} registros, {len(cortes)} cortes")
